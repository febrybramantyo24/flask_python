import shutil
from flask import Flask, Response, redirect, render_template, request, send_file, url_for
from flask import send_file, after_this_request
from flask import send_file
import pandas as pd
import datetime
import os
import sqlalchemy as db
from sqlalchemy import create_engine, text
import glob
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
import sys
import pyautogui
from openpyxl.styles import numbers
import time


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    # Dapatkan input dari form pada halaman web
    disbursement_id = request.form['disbursement_id']
    selisih_tanggal = request.form['selisih_tanggal']
    user_input = request.form['user_input']

    # Validasi Disbursement ID harus berupa angka
    if not disbursement_id.isdigit():
        return render_template('error.html', message="Disbursement ID harus berupa angka.")

    # Validasi Selisih Tanggal harus berisi angka dan karakter '-' atau '+'
    #if not selisih_tanggal.isdigit():
    #    return render_template('error.html', message="Selisih Tanggal harus berisi angka.")

    
    # Mengakses file yang diunggah
    file = request.files['file']
    
    # Memeriksa apakah ada file yang diunggah
    if file:
            # Memeriksa ekstensi file
        if file.filename.endswith('.xls') or file.filename.endswith('.xlsx'):
            # Menyimpan file yang diunggah ke direktori data_inject
            folder_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'data_inject'))
        # Membuat direktori "data_inject" jika belum ada
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            
            # Menyimpan file yang diunggah ke direktori "data_inject"
            file.save(os.path.join(folder_path, file.filename))
            latest_file = max(glob.glob(os.path.join(folder_path, '*.xlsx')), key=os.path.getctime)

            df = pd.read_excel(os.path.join(folder_path, latest_file), sheet_name=0)

            wb = load_workbook(os.path.join(folder_path, file.filename))

            #------ delete 3 sheet 
            if "copy_schedule" in wb.sheetnames:
                sheet_to_delete = wb["copy_schedule"]
                wb.remove(sheet_to_delete)

            if "db_before" in wb.sheetnames:
                sheet_to_delete = wb["db_before"]
                wb.remove(sheet_to_delete)

            if "csv" in wb.sheetnames:
                sheet_to_delete = wb["csv"]
                wb.remove(sheet_to_delete)

            #-------"sheet copy_schedule"
            sheet_name = 'copy_schedule'
            if sheet_name in wb.sheetnames:
                print(f'Sheet {sheet_name} already exists in {latest_file}')
            else:
                new_sheet = wb.create_sheet(sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    new_sheet.append(r)

            ws_copy_schedule = wb['copy_schedule'] # Convert datetime yyyy-mm-dd
            for row in range(2, ws_copy_schedule.max_row + 1):
                date_str = str(ws_copy_schedule.cell(row=row, column=4).value)
                if date_str == 'NaT':
                    continue  # skip this row
                date_only_str = date_str.split()[0]
                date_obj = datetime.strptime(date_only_str, '%Y-%m-%d')
                new_date_str = date_obj.strftime('%Y-%m-%d')
                ws_copy_schedule.cell(row=row, column=4).value = new_date_str

            # Looping untuk menghapus karakter ".0" jika ada
            for row in range(2, ws_copy_schedule.max_row + 1):
                cell_value = ws_copy_schedule.cell(row=row, column=3).value
                if cell_value is not None and isinstance(cell_value, str) and cell_value.endswith('.0'):
                    ws_copy_schedule.cell(row=row, column=3).value = cell_value.rstrip('.0')

            wb.save(os.path.join(folder_path, latest_file))

            #------Connection To Database---
            engine = db.create_engine('mysql+pymysql://mekar:d1saa%40tr3v0_QuK3but@rm-d9jbw9r6mim1jv8n2wo.mysql.ap-southeast-5.rds.aliyuncs.com/stg_funderportal') # Set up database connection
            connection = engine.connect()

            query = f"""
            SELECT
                repaymentschedule.id AS repaymentschedule_id, 
                partner.name AS partner_name,
                instrument.display_name AS display_name,
                repaymentschedule.repayment_schedule_date,
                repaymentschedule.interest,
                repaymentschedule.principal_amount,
                repaymentschedule.amount,
                repaymentschedule.tax_amount
            FROM
                payment_repaymentschedule repaymentschedule
                JOIN 
                instrument_instrument instrument ON instrument.id = repaymentschedule.instrument_id
                JOIN 
                instrument_partner partner ON partner.id = instrument.partner_id
            WHERE
                disbursement_id = {disbursement_id} 
            ORDER BY repaymentschedule.created_date;
            """
            
            db_before = pd.read_sql_query(query, connection)
            print(f"{db_before}")

            #----
            new_sheet_db_before = wb.create_sheet("db_before")

            for r in dataframe_to_rows(db_before, index=False, header=True):
                new_sheet_db_before.append(r)

            sheet_db_before = 'db_before'
            new_sheet_db_before = wb[sheet_db_before]
            print(f'Sheet {sheet_db_before} created and data copied successfully to {new_sheet_db_before.title}.')

            wb.save(os.path.join(folder_path, latest_file))
            connection.close()

            #-----"sheet db_before" 
            ws_copy_schedule = wb['copy_schedule']
            ws_db_before = wb['db_before']
            selisih = timedelta(days=int(selisih_tanggal))

            for row in ws_db_before.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (datetime, date)):
                        cell_value_str = cell.value.strftime('%d/%m/%Y')
                        cell_date = datetime.strptime(cell_value_str, '%d/%m/%Y')
                        cell_date += selisih
                        cell.value = cell_date

            print (ws_db_before.columns)        

            #----- cek apa ada perbedaan displayname 0 di depan
            if user_input == "":
                print("Skip - Tidak ada input dari pengguna")
            elif user_input == "0":
                for row in ws_db_before.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell_value = str(cell.value)
                        if cell_value.startswith('0'):
                            cell.value = cell_value[1:]
                        else:
                            continue  # Skip this row if no '0' at the beginning
            elif user_input == "00":
                for row in ws_db_before.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell_value = str(cell.value)
                        if cell_value.startswith('00'):
                            cell.value = cell_value[2:]
                        else:
                            continue  # Skip this row if no '00' at the beginning
            elif user_input == "000":
                for row in ws_db_before.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell_value = str(cell.value)
                        if cell_value.startswith('000'):
                            cell.value = cell_value[3:]
                        else:
                            continue  # Skip this row if no '000' at the beginning
            else:
                print("Input tidak valid - Skip")


            wb.save(os.path.join(folder_path, latest_file))

            for row in range(2, ws_db_before.max_row + 1):
                date_str = str(ws_db_before.cell(row=row, column=4).value)
                date_only_str = date_str.split()[0]  
                date_obj = datetime.strptime(date_only_str, '%Y-%m-%d')
                new_date_str = date_obj.strftime('%Y-%m-%d')
                ws_db_before.cell(row=row, column=4).value = new_date_str

            wb.save(os.path.join(folder_path, latest_file))

            # ================= test 
            for row in range(2, ws_copy_schedule.max_row + 1):
                cell_value = ws_copy_schedule.cell(row=row, column=3).value
                if isinstance(cell_value, float) and cell_value.is_integer():
                    ws_copy_schedule.cell(row=row, column=3).value = str(int(cell_value))
                else:
                    continue

                cell = ws_copy_schedule.cell(row=row, column=3)
                cell.number_format = numbers.FORMAT_TEXT
            #---------------

            #print (ws_db_before.columns)
            wb.save(os.path.join(folder_path, latest_file))

            #------"CONCATENATE sheet db_before"

            for row in range(2, ws_db_before.max_row + 1):
                cell_c = ws_db_before[f'C{row}'].value
                cell_d = ws_db_before[f'D{row}'].value
                
                concat_value = f'{cell_c}{cell_d}'
                ws_db_before[f'J{row}'].value = concat_value
                
                # convert the formula in cell J to its value
                cell_j = ws_db_before[f'J{row}']
                if cell_j.data_type == 'f':  
                    cell_j.value = cell_j.value
                
                # copy the value of column A to column K in the current row
                ws_db_before[f'K{row}'].value = ws_db_before[f'A{row}'].value

            print(f'CONCATENATE db_before {cell_j}')

            #------"CONCATENATE sheet copy_schedule"
                    
            for row in range(2, ws_copy_schedule.max_row + 1):
                # get the values of cells C and D in the current row
                cell_c = ws_copy_schedule[f'C{row}'].value
                cell_d = ws_copy_schedule[f'D{row}'].value

                if cell_c and cell_d:
                    concat_value = f'{cell_c}{cell_d}'
                    ws_copy_schedule[f'J{row}'].value = concat_value
                else:
                    ws_copy_schedule[f'J{row}'].value = "Data not found"

                cell_j = ws_copy_schedule[f'J{row}']
                if cell_j.data_type == 'f':  # check if the cell contains a formula
                    cell_j.value = cell_j.value

            print(f'CONCATENATE copy_schedule {cell_j}')
            wb.save(os.path.join(folder_path, latest_file))

            max_row_copy = ws_copy_schedule.max_row
            last_row = ws_db_before.max_row

            db_dict = {}
            for row in ws_db_before.iter_rows(min_row=2, min_col=10, max_col=11):
                j_value = row[0].value
                k_value = row[1].value
                db_dict[j_value] = k_value

            for row in ws_copy_schedule.iter_rows(min_row=2, min_col=9, max_col=10):
                j_value = row[1].value
                if j_value in db_dict:
                    k_value = db_dict[j_value]
                else:
                    k_value = None
                ws_copy_schedule.cell(row=row[0].row, column=11, value=k_value)

                b_value = ws_db_before.cell(row=row[0].row, column=2).value
                ws_copy_schedule.cell(row=row[0].row, column=2, value=b_value)

            ws_copy_schedule.cell(row=1, column=11).value = 'VLOOKUP_RESULT'
            ws_copy_schedule.cell(row=1, column=10).value = 'CONCATENATE_RESULT'

            # Looping untuk menghapus karakter ".0" jika ada
            for row in range(2, ws_db_before.max_row + 1):
                cell_value = ws_db_before.cell(row=row, column=3).value
                if cell_value is not None and isinstance(cell_value, str) and cell_value.endswith('.0'):
                    ws_db_before.cell(row=row, column=3).value = cell_value.rstrip('.0')

            for row in ws_copy_schedule.iter_rows(min_row=2, min_col=11, max_col=11):
                for cell in row:
                    ws_copy_schedule.cell(row=cell.row, column=1).value = cell.value

            for row in ws_copy_schedule.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell_value = cell.value
                    if isinstance(cell_value, float) and cell_value.is_integer():
                        cell_value = int(cell_value)
                    ws_copy_schedule.cell(row=cell.row, column=3).value = str(cell_value).rstrip('.0')

            if k_value is None:
                print("Check : Masih terdapat #N/A")
            else:
                print(f"Berhasil menemukan repaymentschedule_id {k_value}")

            wb.save(os.path.join(folder_path, latest_file))

            #-------
            if 'csv' in wb.sheetnames:
                ws_csv = wb['csv']
                wb.remove(ws_csv)

            ws_csv = wb.create_sheet('csv')

            # Mengambil header dari sheet "copy_schedule"
            header_row = []
            for cell in ws_copy_schedule[1]:
                header_row.append(cell.value)

            # Menambahkan header ke sheet "csv"
            ws_csv.append(header_row)

            # Mengambil data dari sheet "copy_schedule" dan menyortir berdasarkan kolom A
            data_to_copy = []
            for row in ws_copy_schedule.iter_rows(min_row=2, min_col=1, max_col=7):
                if row[0].value is not None:
                    data_to_copy.append([cell.value for cell in row])

            data_to_copy.sort(key=lambda x: x[0])
            # Menghapus konten dari sheet "csv"
            ws_csv.delete_rows(2, ws_csv.max_row)

            # Menyalin data yang telah disortir ke sheet "csv"
            for row_index, data_row in enumerate(data_to_copy, start=2):
                for col_index, value in enumerate(data_row, start=1):
                    ws_csv.cell(row=row_index, column=col_index, value=value)

            ws_csv.delete_cols(8, 13)

            wb.save(os.path.join(folder_path, latest_file))
            #return 'Data processed successfully.'
    df_export = pd.read_excel(os.path.join(folder_path, latest_file), sheet_name='csv')
    # Mengembalikan file CSV dan XLSX kepada pengguna untuk diunduh
    result = "Data hasil proses: " + str(df_export)
    return render_template('result.html', result=result)
            


@app.route('/download/<file_type>', methods=['GET'])
def download(file_type):
    folder_path = os.path.join(os.path.dirname(__file__), "data_inject")
    files = glob.glob(os.path.join(folder_path, '*'))

    if file_type == 'csv':
        # Membuat file hasil proses dalam format CSV
        csv_file = os.path.join(folder_path, 'hasil_proses2.csv')
        df = pd.read_excel(files[0], sheet_name='csv')
        if os.path.exists(csv_file):
            existing_df = pd.read_csv(csv_file)
            df = pd.concat([existing_df, df])
        df.to_csv(csv_file, index=False)
        return send_file(csv_file, as_attachment=True)
    
    elif file_type == 'xlsx':
            # Mengambil file XLSX terbaru dari folder "data_inject"
            xlsx_files = [file for file in files if file.endswith('.xlsx')]
            if xlsx_files:
                latest_xlsx_file = max(xlsx_files, key=os.path.getctime)
                return send_file(latest_xlsx_file, as_attachment=True)
            else:
                return render_template('error.html', error_message='No XLSX files found in the "data_inject" folder.')

    else:
        return render_template('error.html', error_message='Invalid file type.')


@app.route('/delete_files', methods=['POST'])
def delete_files():
    folder_path = os.path.join(os.path.dirname(__file__), "data_inject")
    # Menghapus semua file di folder "data_inject"
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)
    return render_template('index.html')

    
if __name__ == '__main__':
    app.run(debug=True)
